# 文件路径: src/excel_exporter.py
# 最终优化版: 增加了打印时自动缩放以适应页面宽度的功能

import os
import datetime
import logging
import pandas as pd
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

class ExcelExporter:
    def __init__(self, config_manager):
        self.config_manager = config_manager

    def get_output_dir(self):
        output_dir = self.config_manager.get("excel_output_dir")
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except OSError as e:
                logging.error(f"无法创建Excel输出目录 {output_dir}: {e}")
                return None
        return output_dir

    def _to_chinese_currency(self, num):
        cn_num = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
        cn_unit = ['', '拾', '佰', '仟', '万', '拾万', '佰万', '仟万', '亿', '拾亿', '佰亿', '仟亿', '兆']
        cn_decimal = ['角', '分']
        num_str = f"{num:.2f}"
        integer_part, decimal_part = num_str.split('.')
        integer_result = ""
        if integer_part == '0': integer_result = '零'
        else:
            integer_len = len(integer_part)
            for i in range(integer_len):
                digit = int(integer_part[i])
                unit = cn_unit[integer_len - 1 - i]
                if digit == 0:
                    if integer_result and integer_result[-1] != '零': integer_result += '零'
                else: integer_result += cn_num[digit] + unit
            if integer_result.endswith('零'): integer_result = integer_result[:-1]
        decimal_result = ""
        if decimal_part == '00': decimal_result = '整'
        else:
            if decimal_part[0] != '0': decimal_result += cn_num[int(decimal_part[0])] + cn_decimal[0]
            if decimal_part[1] != '0': decimal_result += cn_num[int(decimal_part[1])] + cn_decimal[1]
        if integer_part == '0': return f"零元{decimal_result}"
        return f"{integer_result}元{decimal_result}"

    def create_settlement_workbook(self, df, title, entity_name, record_type, date_range=None):
        company_name = self.config_manager.get("company_name", "公司名称")
        phone_number = self.config_manager.get("phone_number", "")
        footer_text = self.config_manager.get("footer_text", "")
        
        df_processed = df.copy()
        df_processed.rename(columns={'grower_name': '姓名', 'client_name': '姓名', 'date': '日期', 'spec': '规格', 'unit_price': '单价', 'total_amount': '金额', 'notes': '备注'}, inplace=True)
        
        if record_type == 'grower':
            df_processed.rename(columns={'net_weight': '净重(斤)', 'gross_weight':'毛重(斤)', 'secondary_fruit':'次果(斤)', 'tare_weight':'皮重(斤)'}, inplace=True)
            display_columns = ['日期', '规格', '毛重(斤)', '次果(斤)', '皮重(斤)', '净重(斤)', '单价', '金额', '备注']
        else: # client
            df_processed.rename(columns={'pieces': '件数', 'weight': '重量(斤)'}, inplace=True)
            display_columns = ['日期', '规格', '件数', '重量(斤)', '单价', '金额', '备注']

        for col in ['净重(斤)', '毛重(斤)', '次果(斤)', '皮重(斤)', '件数', '重量(斤)', '单价', '金额']:
            if col in df_processed.columns:
                df_processed[col] = pd.to_numeric(df_processed[col], errors='coerce').fillna(0)
        
        df_display = df_processed[display_columns]
        total_amount = df_display['金额'].sum()

        font_company = Font(name='黑体', size=20, bold=True, color="002060")
        font_title = Font(name='宋体', size=16, bold=True)
        font_info = Font(name='宋体', size=11)
        font_header = Font(name='宋体', size=11, bold=True, color="FFFFFF")
        font_body = Font(name='宋体', size=10)
        font_total_label = Font(name='宋体', size=11, bold=True)
        font_total_value = Font(name='宋体', size=11, bold=True, color="FF0000")
        font_footer = Font(name='宋体', size=9, italic=True, color="808080")
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        fill_row_alt = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        border_thin_side = Side(border_style="thin", color="BFBFBF")
        border_thin_box = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
        currency_format = '¥#,##0.00'
        number_format = '#,##0.00'
        
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "结算明细"
        
        num_cols = len(display_columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        cell = ws.cell(1, 1, company_name)
        cell.font = font_company
        cell.alignment = align_center
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        cell = ws.cell(2, 1, title)
        cell.font = font_title
        cell.alignment = align_center
        ws.row_dimensions[2].height = 25
        
        date_range_str = f"{date_range[0]} 至 {date_range[1]}" if date_range and date_range[0] != date_range[1] else (date_range[0] if date_range else "")
        ws.cell(4, 1, f"结算对象: {entity_name}").font = font_info
        ws.merge_cells(start_row=4, start_column=num_cols - 2, end_row=4, end_column=num_cols)
        cell = ws.cell(4, num_cols - 2, f"结算周期: {date_range_str}")
        cell.font = font_info
        cell.alignment = align_right
        
        header_row = 6
        for c_idx, col_name in enumerate(display_columns, 1):
            cell = ws.cell(header_row, c_idx, col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin_box
        ws.row_dimensions[header_row].height = 20
        
        current_row = header_row + 1
        for r_idx, row_data in enumerate(df_display.itertuples(index=False), 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(current_row, c_idx, value)
                cell.font = font_body
                cell.border = border_thin_box
                col_name = display_columns[c_idx-1]
                if col_name in ['净重(斤)', '毛重(斤)', '次果(斤)', '皮重(斤)', '件数', '重量(斤)', '单价', '金额']:
                    cell.alignment = align_right
                    if col_name in ['单价', '金额']: cell.number_format = currency_format
                    else: cell.number_format = number_format
                else:
                    cell.alignment = align_left
                if r_idx % 2 == 0: cell.fill = fill_row_alt
            current_row += 1
            
        total_row_idx = current_row + 1
        total_amount_col_idx = display_columns.index('金额') + 1
        ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=total_amount_col_idx - 1)
        cell = ws.cell(total_row_idx, 1, "总计 (TOTAL)")
        cell.font = font_total_label
        cell.alignment = align_right
        
        cell = ws.cell(total_row_idx, total_amount_col_idx, total_amount)
        cell.font = font_total_value
        cell.alignment = align_right
        cell.number_format = currency_format
        
        cn_total_amount = self._to_chinese_currency(total_amount)
        amount_in_words_row = total_row_idx + 1
        ws.merge_cells(start_row=amount_in_words_row, start_column=1, end_row=amount_in_words_row, end_column=num_cols)
        cell = ws.cell(amount_in_words_row, 1, f"金额大写: {cn_total_amount}")
        cell.font = font_info
        cell.alignment = align_left
        
        footer_row_start = amount_in_words_row + 2
        ws.merge_cells(start_row=footer_row_start, start_column=1, end_row=footer_row_start, end_column=num_cols)
        ws.cell(footer_row_start, 1, footer_text).font = font_footer
        ws.merge_cells(start_row=footer_row_start + 1, start_column=1, end_row=footer_row_start + 1, end_column=num_cols)
        cell = ws.cell(footer_row_start + 1, 1, f"联系电话: {phone_number} | 生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        cell.font = font_footer
        cell.alignment = align_right
        
        fixed_widths = {'日期': 12, '单价': 10, '金额': 15}
        
        for idx, col_name in enumerate(display_columns, 1):
            column_letter = get_column_letter(idx)
            if col_name in fixed_widths:
                ws.column_dimensions[column_letter].width = fixed_widths[col_name]
            else:
                max_length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in str(col_name))
                for cell in ws[column_letter]:
                    if cell.value is not None:
                        try:
                            cell_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in str(cell.value))
                            if cell_len > max_length: max_length = cell_len
                        except: pass
                adjusted_width = max_length + 2
                if col_name == '备注': adjusted_width = min(adjusted_width, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        # --- 核心修改点：添加打印设置 ---
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1  # <-- 关键设置：fitToWidth = 1 表示缩放为1页宽
        ws.page_setup.fitToHeight = 0 # <-- fitToHeight = 0 表示高度可以自动延伸
        
        ws.print_title_rows = f'{header_row}:{header_row}'
        ws.page_margins.left = 0.75; ws.page_margins.right = 0.75; ws.page_margins.top = 0.75; ws.page_margins.bottom = 0.75
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True
        
        return wb, entity_name

    def save_and_notify(self, wb, entity_name, title):
        output_dir = self.get_output_dir()
        if not output_dir: return

        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        safe_entity_name = "".join(c for c in entity_name if c.isalnum() or c in (' ', '_')).rstrip()
        file_name = f"{safe_entity_name}_{title}_{timestamp}.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        try:
            wb.save(file_path)
            messagebox.showinfo("导出成功", f"{title}已成功导出到:\n{file_path}")
            os.startfile(output_dir)
        except PermissionError:
            messagebox.showerror("导出失败", f"文件权限不足或文件被占用，无法写入：\n{file_path}\n请关闭可能正在使用此文件的Excel程序。")
        except Exception as e:
            logging.error(f"导出Excel时发生未知错误: {e}", exc_info=True)
            messagebox.showerror("导出失败", f"导出Excel时发生未知错误: {e}\n详情请查看logs/app.log文件。")